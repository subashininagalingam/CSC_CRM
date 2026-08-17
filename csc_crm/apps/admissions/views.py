import os
from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.core.paginator import Paginator
from .forms import StudentForm, AdmissionForm, EnrollmentForm 
from .models import *
from django.contrib import messages
from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate, Table
from .filters import StudentFilter
from django.utils import timezone
from django.db import transaction
from . services import get_fee_summary
from django.conf import settings
from django.http import JsonResponse
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import TableStyle
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from csc_crm.apps.student_attendance.models import Batch
import uuid
from csc_crm.apps.admissions.models import Payment
from django.contrib.staticfiles import finders
from playwright.async_api import async_playwright
from asgiref.sync import async_to_sync
from django.urls import reverse
from django.core.mail import send_mail
from urllib import request
from django.db import models
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required

# ================= STUDENT REGISTRATION ====================

def student(request):

    print(request.POST)
    print("BATCH =", request.POST.get("batch"))
    student_form = StudentForm()
    admission_form = AdmissionForm()
    enrollment_form = EnrollmentForm()
    courses = Course.objects.all()

    if request.method == "POST":

        student_form = StudentForm(request.POST, request.FILES)
        admission_form = AdmissionForm(request.POST)
        enrollment_form = EnrollmentForm(request.POST, request.FILES)

        if (student_form.is_valid() and admission_form.is_valid() and enrollment_form.is_valid()):
            try:
                with transaction.atomic():
                    student = student_form.save()
                    id_proofs = request.FILES.getlist("id_proof")

                    for file in id_proofs:
                         StudentDocument.objects.create(
                         student=student,
                         document_type="id_proof",
                         document=file
                        )
                    certificates = request.FILES.getlist("certificate")
                    
                    for file in certificates:
                         StudentDocument.objects.create(
                         student=student,
                         document_type="certificate",
                         document=file
                       )
                    admission = Admission.objects.create(
                        student=student,
                        course_name=admission_form.cleaned_data['course_name'],
                        status=admission_form.cleaned_data['status']
                        )
                    enrollment = enrollment_form.save(commit=False)

                    if ( enrollment.batch.course_id!= admission.course_name_id):
                        messages.error(
                            request,
                            "Selected batch does not belong to selected course."
                        )
                        raise ValueError( "Batch and Course mismatch")

                    enrollment.admission = admission
                    print("Before save")
                    enrollment.save()
                    print("After save")
                
                    try:
                        # Confirmation emails to student and admin
                        # (currently disabled, left commented out below)

                        # USER EMAIL

#                         send_mail(
#                             subject='🎓 Admission Confirmation - CSC Academy',

#                             message=f'''
# Dear {student.first_name},

# Congratulations! 🎉

# Your admission has been successfully confirmed at CSC Academy.

# Course Enrolled:
# {admission.course_name}

# We are excited to have you as part of our learning journey.

# Best Regards,
# CSC Academy
# ''',

#                             from_email=settings.EMAIL_HOST_USER,

#                             recipient_list=[student.email],

#                             fail_silently=False,
#                         )

                        # ADMIN EMAIL

#                         send_mail(
#                             subject='📌 New Student Admission Alert',

#                             message=f'''
# A new student admission has been registered.

# Student Details
# -------------------------

# Name   : {student.first_name} {student.last_name}

# Course : {admission.course_name}

# Phone  : {student.phone_no}

# Email  : {student.email}

# Please verify the records from the admin panel.
# ''',

#                             from_email=settings.EMAIL_HOST_USER,

#                             recipient_list=['admin@gmail.com'],

#                             fail_silently=False,
#                         )

                        print("MAIL SENT SUCCESS")

                    except Exception as e:
                        print("MAIL ERROR:", e)

                    messages.success(request, "Student enrolled successfully!")

                    return redirect('fee_dashboard')
            except Exception as e:

                print("ERROR:", e)

        else:

            messages.error(request, "Form has errors. Please check!")
            print("FORM ERRORS :", student_form.errors, admission_form.errors, enrollment_form.errors)

    return render(request, 'admissions/register.html', {
        'student_form': student_form,
        'admission_form': admission_form,
        'enrollment_form': enrollment_form,
        'courses': courses,
    })
#=================== STUDENT DOCUMENT DELETE ====================
def delete_student_document(request, pk):
    """Delete a student document"""
    document = get_object_or_404(StudentDocument, pk=pk)
    student_id = document.student.id
    document_display_name = (
        os.path.basename(document.document.name)
        if document.document else "Document"
    )

    if request.method == "POST":
        if document.document:
            document.document.delete(save=False)
        document.delete()
        messages.success(
            request,
            f"Document '{document_display_name}' deleted successfully."
        )

    return redirect("edit_student", id=student_id)

#=================== FEE DASHBOARD ====================

def _get_summary_value(summary, key, default=0):
    """Works whether `summary` is a dict or a plain object/namedtuple."""
    if isinstance(summary, dict):
        return summary.get(key, default)
    return getattr(summary, key, default)

def fee_dashboard(request):
    students = Student.objects.all().order_by('-id')
    fee_students = []
    
    for student in students:    
        if student.pending_amount() > 0:
            fee_students.append(student)
            
    latest_payments = []
    seen_students = set()
    
    for payment in Payment.objects.order_by('-date', '-id'): 
        if payment.student.id not in seen_students:
            latest_payments.append(payment)
            seen_students.add(payment.student.id)
        if len(latest_payments) == 5:
            break
        
    selected_student_id = request.GET.get('student_id')
    remaining_payments = 0

    # SAVE PAYMENT
    if request.method == 'POST':
        student_id = request.POST.get('student')
        amount = request.POST.get('amount')
        try:
            amount = float(amount)
        except:
            messages.error(request, "Invalid amount")
            return redirect('fee_dashboard')

        mode = request.POST.get('mode')
        reference = request.POST.get('reference')
        if not reference or not reference.strip():
            reference = f"TXN{uuid.uuid4().hex[:8].upper()}"
            
        remarks = request.POST.get('remarks')
        student = Student.objects.get(id=student_id)
        total_fee = student.total_fee()
        paid_amount = student.total_paid()
        pending_amount = total_fee - paid_amount
        
        if amount <= 0:
            messages.error(request, "Amount must be greater than 0.")
            return redirect('fee_dashboard')
       
        if amount > pending_amount:
            messages.error(request,f"Only remaining amount ₹{pending_amount} can be paid.")
            return redirect('fee_dashboard')

        payment_count = Payment.objects.filter(student=student).count()

        if payment_count >= 6:
            messages.error(request, "Only 6 payments allowed.")
            return redirect('fee_dashboard')

        if payment_count == 5:
            if amount != pending_amount:
                messages.error(request,f"6th payment must clear full remaining amount ₹{pending_amount}")
                return redirect('fee_dashboard')

        Payment.objects.create(
            student_id=student_id,
            amount=amount,
            mode=mode,
            reference_id=reference,
            remarks=remarks
        )

        remaining_payments = 6 - (payment_count + 1)
        new_pending = pending_amount - amount
        if new_pending <= 0:
            messages.success(request, "Payment Successful! Full fee has been paid.")
        else:
            messages.success(
        request,
        f"Payment Successful! Remaining payments: {remaining_payments}"
    )
    
        return redirect('fee_dashboard')
    
    # ====== EXPORT EXCEL=======

    format = request.GET.get('format')
    if format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Payments"

        ws.append(["Student","Course","Batch","Amount","Mode","Reference","Date"])

        header_fill = PatternFill( start_color="FFC000",end_color="FFC000",fill_type="solid")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for payment in Payment.objects.select_related('student').order_by('-date', '-id'):
            admission = payment.student.admissions.first()
            enrollment = (
            admission.enrollment
            if admission and hasattr(admission, 'enrollment')
            else None
        )

            ws.append([
            f"{payment.student.first_name} {payment.student.last_name}",
            str(admission.course_name) if admission else "-",
            str(enrollment.batch) if enrollment else "-",
            payment.amount,
            payment.mode,
            payment.reference_id,
            str(payment.date),
        ])
        column_widths = {'A': 25,'B': 25,'C': 18,'D': 15,'E': 15,'F': 20,'G': 18,}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = ('attachment; filename=fee_payments.xlsx')
        wb.save(response)
        return response

    # ============DASHBOARD SUMMARY============
    summary = get_fee_summary()  
    collected = _get_summary_value(summary, 'collected', 0) or 0        
    outstanding = _get_summary_value(summary, 'outstanding', 0) or 0    
    total_expected = collected + outstanding                           

    if total_expected > 0:                                              
        collection_percentage = round((collected / total_expected) * 100)  
    else:                                                                
        collection_percentage = 0                                      
    student_fee_status = []

    for student in students:

        total_fee = student.total_fee()
        paid = student.total_paid()
        pending = student.pending_amount()

        if pending <= 0:
            status = 'Paid'
        elif paid == 0:
            status = 'Pending'
        else:
            status = 'Partial'

        student_fee_status.append({
            'student': student,
            'total_fee': total_fee,
            'paid': paid,
            'pending': pending,
            'status': status
        })

    context = {
        'students': fee_students,
        'payments': latest_payments,
        'summary': summary,
        'remaining_payments': remaining_payments,
        'student_fee_status': student_fee_status,
        'selected_student_id': selected_student_id,
        'collection_percentage': collection_percentage,   
    }

    return render(
        request,
        'admissions/fee_dashboard.html',
        context
    )

#=====================STUDENT DETAILS======================

def student_detail(request, pk):
    student = Student.objects.get(id=pk)
    payments = student.payments.all()

    context = {
        'student': student,
        'payments': payments,
        'total_paid': student.total_paid(),
        'pending': student.pending_amount(),
    }

    return render(request, 'students/detail.html', context)
#==============LINK CALLBACK FOR PDF GENERATION================

def link_callback(uri):
    if uri.startswith(settings.STATIC_URL):
        path = finders.find( uri.replace(settings.STATIC_URL, ""))
        if path:
         return path
    return uri
async def html_to_pdf(url):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.goto(url, wait_until="networkidle")
        await page.wait_for_load_state("networkidle")
        await page.emulate_media(media="print")

        pdf = await page.pdf(
            format="A4",
            print_background=True,
            prefer_css_page_size=True,
            margin={ "top": "8mm", "bottom": "8mm", "left": "8mm", "right": "8mm",},
        )
        await browser.close()
        return pdf
    
    #=================== PDF RECEIPT GENERATION ====================
    
def generate_receipt(request, pk):

    payment = Payment.objects.get(id=pk)
    admission = payment.student.admissions.first()
    enrollment = (
        admission.enrollment
        if admission and hasattr(admission, 'enrollment')
        else None
    )

    context = {
        'payment': payment,
        'student_name':
        f"{payment.student.first_name} {payment.student.last_name}",
        'phone':
        payment.student.phone_no,
        'email':
        payment.student.email,
        'course':
        admission.course_name if admission else "-",
        'batch':
        enrollment.batch if enrollment else "-",
        'total_fee':
        payment.student.total_fee(),
        'total_paid':
        payment.student.total_paid(),
        'pending':
        payment.student.pending_amount(),
    }

    url = request.build_absolute_uri( reverse("receipt_pdf", args=[pk]) )
    pdf = async_to_sync(html_to_pdf)(url)
    response = HttpResponse( pdf,content_type="application/pdf")
    response["Content-Disposition"] = ( f'attachment; filename="receipt_{pk}.pdf"')

    return response

#=================== STUDENT LIST ====================

@login_required(login_url='staff_login')
def student_list(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    students = Student.objects.prefetch_related(
        'payments',
        'admissions__enrollment',
        'admissions__course_name',
    ).all()

    # Trainer -> only students from own assigned batches
    if my_role == 'Trainer':
        students = students.filter(
            admissions__enrollment__batch__trainer=my_staff
        ).distinct()

    student_filter = StudentFilter(request.GET, queryset=students)
    filtered_students = student_filter.qs.distinct().order_by('-id')
    paginator = Paginator(filtered_students, 10)  # Show 10 students per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    active_students_qs = Admission.objects.filter(status="enrolled")
    if my_role == 'Trainer':
        active_students_qs = active_students_qs.filter(enrollment__batch__trainer=my_staff)
    active_students = active_students_qs.count()

    summary = get_fee_summary()
    format = request.GET.get('format')

    if format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = ["Student", "Course", "Batch", "Phone", "Payment Status", "Joined"]
        ws.append(headers)
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for s in filtered_students:
            for admission in s.admissions.all():
                enrollment = getattr(admission, 'enrollment', None)
                ws.append([
                    f"{s.first_name} {s.last_name}",
                    str(admission.course_name) if admission else "-",
                    str(enrollment.batch) if enrollment and enrollment.batch else "-",
                    s.phone_no,
                    enrollment.payment_status if enrollment else "-",
                    str(enrollment.start_date) if enrollment else "-"
                ])
        column_widths = {'A': 25,'B': 20,'C': 15,'D': 18,'E': 18,'F': 18,}

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        wb.save(response)
        return response
    # ================= PDF =================
    
    if format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="students.pdf"'
        data = [["Student", "Course", "Batch", "Phone", "Payment Status", "Joined"]]

        for s in filtered_students:
            for admission in s.admissions.all():
                enrollment = getattr(admission, 'enrollment', None)

                data.append([
                    f"{s.first_name} {s.last_name}",
                    str(admission.course_name) if admission else "-",
                    str(enrollment.batch) if enrollment and enrollment.batch else "-",
                    s.phone_no,
                    enrollment.payment_status if enrollment else "-",
                    str(enrollment.start_date) if enrollment else "-"
                ])

        doc = SimpleDocTemplate(response)
        styles = getSampleStyleSheet()
        title = Paragraph("Student Report", styles['Title'])
        table = Table(data)
        table.setStyle(TableStyle([ ('BACKGROUND', (0, 0), (-1, 0), colors.gold), ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),]))
        elements = [ title, Spacer(1, 12), table]
        doc.build(elements)

        return response

    batches_qs = Batch.objects.all()
    if my_role == 'Trainer':
        batches_qs = batches_qs.filter(trainer=my_staff)

    return render(request, 'admissions/student_list.html', {
        'page_obj': page_obj,
        'filter': student_filter,
        'courses': Course.objects.all(),
        'batches': batches_qs,
        'total_students': filtered_students.count(),
        'active_students': active_students,
        'summary': summary
    })

#=================== EDIT STUDENT ====================

def edit_student(request, id):

    student = get_object_or_404(Student, id=id)
    admission = Admission.objects.get(student=student)
    enrollment = Enrollment.objects.get(admission=admission)
    courses = Course.objects.all()
    batches = Batch.objects.filter(course_id=admission.course_name_id) if admission.course_name_id else Batch.objects.none()

    if request.method == 'POST':

        student.first_name = request.POST.get('first_name')
        student.last_name = request.POST.get('last_name')
        student.email = request.POST.get('email')
        student.phone_no = request.POST.get('phone_no')
        student.dob = request.POST.get('dob')
        student.gender = request.POST.get('gender')
        student.guardian_name = request.POST.get('guardian_name')
        student.guardian_phone_no = request.POST.get('guardian_phone_no')
        student.address = request.POST.get('address')

        if request.FILES.get('photo'):
            student.photo = request.FILES.get('photo')

        if request.FILES.get('id_proof'):
            student.id_proof = request.FILES.get('id_proof')

        if request.FILES.get('certificate'):
            student.certificate = request.FILES.get('certificate')

        student.save() 

        course_id = request.POST.get('course_name')
        if course_id:
            admission.course_name_id = course_id
        admission.status = request.POST.get('status')
        admission.save()

        batch_id = request.POST.get('batch')

        if batch_id:
            enrollment.batch_id = batch_id

        enrollment.start_date = request.POST.get('start_date')
        enrollment.save()
        messages.success(request, "✅ Student details updated successfully")
        return redirect('student_list')

    context = {
        'student': student,
        'admission': admission,
        'enrollment': enrollment,
        'courses': courses,
        'batches' : batches
    }
    return render(request, "admissions/edit_student.html", context)

#=================== DELETE STUDENT ====================
def delete_student(request, id):
    student=Student.objects.get(id=id)
    student.delete()
    messages.success(request, "✅Student deleted successfully")
    return redirect('student_list')

#=================== SEARCH STUDENTS ====================

@login_required(login_url='staff_login')
def search_students(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    students = Student.objects.prefetch_related(
        'admissions__course_name',
        'admissions__enrollment'
    )

    # Trainer -> only students from own assigned batches
    if my_role == 'Trainer':
        students = students.filter(
            admissions__enrollment__batch__trainer=my_staff
        )

    student_filter = StudentFilter(request.GET, queryset=students)
    filtered_students = student_filter.qs.distinct().order_by('-id')

    format = request.GET.get('format')

    if format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"

        headers = ["Student", "Student ID", "Course", "Status", "Fee", "Joined"]
        ws.append(headers)

        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ids_param = request.GET.get('ids')

        if ids_param:
            selected_ids = [i for i in ids_param.split(',') if i.strip().isdigit()]
            export_students = filtered_students.filter(id__in=selected_ids)
        else:
            export_students = filtered_students

        for student in export_students:
            for admission in student.admissions.all():
                enrollment = getattr(admission, 'enrollment', None)

                if student.pending_amount() <= 0:
                    fee_status = "Paid"
                elif student.total_paid() > 0:
                    fee_status = "Partial"
                else:
                    fee_status = "Pending"

                ws.append([
                    f"{student.first_name} {student.last_name}",
                    f"STU{student.id}",
                    str(admission.course_name) if admission.course_name else "-",
                    admission.status,
                    fee_status,
                    str(enrollment.start_date) if enrollment else "-",
                ])

        column_widths = {'A': 25, 'B': 12, 'C': 20, 'D': 15, 'E': 12, 'F': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        response = HttpResponse( content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=search_results.xlsx'
        wb.save(response)
        return response

    #PAGINATION 
    try:
        per_page = int(request.GET.get('per_page', 10))
        if per_page not in (10, 25, 50, 100):
            per_page = 10
    except (TypeError, ValueError):
        per_page = 10

    paginator = Paginator(filtered_students, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    querydict = request.GET.copy()
    querydict.pop('page', None)
    querydict.pop('format', None)
    base_query = querydict.urlencode()

    batches_qs = Batch.objects.all()
    if my_role == 'Trainer':
        batches_qs = batches_qs.filter(trainer=my_staff)

    return render(request, "admissions/search_students.html", {
        'filter': student_filter,
        'page_obj': page_obj,
        'courses': Course.objects.all(),
        'batches': batches_qs,
        'per_page': per_page,
        'base_query': base_query,
    })
#=================== VIEW STUDENT ====================
def view_student(request, id):

    student = Student.objects.prefetch_related('payments','admissions__course_name', 'admissions__enrollment' ).get(id=id)
    admission = student.admissions.first()
    enrollment = admission.enrollment if admission else None
    payments = student.payments.all().order_by('-date')
    return render(request, "admissions/view_student.html", {
        'student': student,
        'admission': admission,
        'enrollment': enrollment,
        'payments': payments,
        'total_paid': student.total_paid(),
        'pending': student.pending_amount(),
    })

#=================== CHECK EMAIL AND PHONE ====================
def check_email(request):

    email = request.GET.get('email')
    exclude_id = request.GET.get('exclude_id')
    qs = Student.objects.filter( email=email)

    if exclude_id:
        qs = qs.exclude(pk=exclude_id)

    exists = qs.exists()

    return JsonResponse({
        'exists': exists
    })
    
def check_phone(request):

    phone = request.GET.get('phone')
    exclude_id = request.GET.get('exclude_id')
    qs = Student.objects.filter( phone_no=phone)
    guardian_qs = Student.objects.filter( guardian_phone_no=phone)

    if exclude_id:
        qs = qs.exclude(pk=exclude_id)
        guardian_qs = guardian_qs.exclude(pk=exclude_id)

    exists = qs.exists()
    guardian_exists = guardian_qs.exists()

    return JsonResponse({
        'exists': exists,
        'guardian_exists': guardian_exists
    })

#=================== PREVIEW RECEIPT ====================    
def preview_receipt(request, pk):

    payment = Payment.objects.get(id=pk)
    admission = payment.student.admissions.first()
    enrollment = (
        admission.enrollment
        if admission and hasattr(admission, 'enrollment')
        else None
    )
    context = {
        'payment': payment,
        'student_name': f"{payment.student.first_name} {payment.student.last_name}",
        'phone': payment.student.phone_no,
        'email': payment.student.email,
        'course': admission.course_name if admission else "-",
        'batch': enrollment.batch if enrollment else "-",
        'total_fee': payment.student.total_fee(),
        'total_paid': payment.student.total_paid(),
        'pending': payment.student.pending_amount(),
        'today': timezone.localdate(),
    }

    return render(
        request,
        "admissions/preview_receipt.html",
        context
    )
  #=================== RECEIPT PDF VIEW ====================    
def receipt_pdf(request, pk):

    payment = Payment.objects.get(id=pk)
    admission = payment.student.admissions.first()
    enrollment = (
        admission.enrollment
        if admission and hasattr(admission, "enrollment")
        else None
    )
    context = {
        "payment": payment,
        "student_name": f"{payment.student.first_name} {payment.student.last_name}",
        "phone": payment.student.phone_no,
        "email": payment.student.email,
        "course": admission.course_name if admission else "-",
        "batch": enrollment.batch if enrollment else "-",
        "total_fee": payment.student.total_fee(),
        "total_paid": payment.student.total_paid(),
        "pending": payment.student.pending_amount(),
        "today": timezone.localdate(),
    }
    return render(
        request,
        "admissions/fee_receipt.html",
        context
    )