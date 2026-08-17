from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

#================ EXCEL STYLE CONSTANTS (FILLS, FONTS, BORDERS) ==================#
class ExcelStyles:
    title_fill = PatternFill("solid", fgColor="1E3A8A")
    header_fill = PatternFill("solid", fgColor="2563EB")
    section_fill = PatternFill("solid", fgColor="059669")
    excellent_fill = PatternFill("solid", fgColor="00B050")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    warning_fill = PatternFill("solid", fgColor="FFEB9C")
    critical_fill = PatternFill("solid", fgColor="FFC7CE")
    white_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

#================ STATUS -> CELL FILL COLOR MAP ==================#
STATUS_FILL_MAP = {
    "Excellent": ExcelStyles.excellent_fill,
    "Good": ExcelStyles.good_fill,
    "Warning": ExcelStyles.warning_fill,
    "Critical": ExcelStyles.critical_fill,
}

#================ EXCEL HEADER ROW STYLER ==================#
def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = ExcelStyles.header_fill
        cell.font = ExcelStyles.white_font
        cell.alignment = ExcelStyles.center
        cell.border = ExcelStyles.thin_border