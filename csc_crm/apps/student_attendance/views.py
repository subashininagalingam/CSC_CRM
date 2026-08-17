import json
from django.conf import settings
from django.contrib import messages
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Font, PatternFill)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table)
from rest_framework import filters, viewsets
from rest_framework.decorators import api_view
from rest_framework.response import Response
from csc_crm.apps.admissions.models import Course, Enrollment
from csc_crm.apps.student_attendance.forms import BatchForm
from .filters import AttendanceFilter
from .models import *
from .serializers import *
from .services import (get_absent_tracker_data, get_low_attendance_data)
from datetime import datetime
from django.template.loader import render_to_string
from .attendance_utils import (
    get_attendance_status,
    get_attendance_counts,
    count_by_status,
    get_student_attendance_percentage,
    get_formatted_latest_attendance_date,
    compute_batch_stats,
    get_status_summary_counts,
    get_batch_analytics_rows,
    get_trainer_display_name,
    get_batch_timing_display,
    filter_batches_by_search,
    build_attendance_email,
    build_attendance_sms,
    build_export_rows,
    STATUS_COLOR_HEX,
)
from .pdf_utils import get_default_table_style
from .excel_utils import ExcelStyles, STATUS_FILL_MAP, style_header_row
from django.contrib.auth.decorators import login_required


#===================================== BATCH CRUD API ====================================#
class BatchViewSet(viewsets.ModelViewSet):

    queryset = Batch.objects.all()
    serializer_class = BatchSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['batch_name', 'timing', 'trainer__trainer_name', 'course__course_name']

    def create(self, request, *args, **kwargs):
        print(request.data)
        return super().create(request, *args, **kwargs)

#========================== MARK ATTENDANCE - BATCH DROPDOWN HELPERS =======================#
def get_batches_by_course(request):
    course_id = request.GET.get('course_id')
    batches = Batch.objects.filter(course_id=course_id)
    data = [{"id": b.id, "batch_name": b.batch_name}
        for b in batches
        if b.student_count < 30
    ]

    return JsonResponse(data, safe=False)

#================================= COURSE DURATION AUTOFILL ====================================#
def get_course_duration(request, course_id):

    try:
        course = Course.objects.get(id=course_id)
        return JsonResponse({"duration": course.duration})

    except Course.DoesNotExist:
        return JsonResponse({"error": "Course not found"}, status=404)

#=============================== DASHBOARD (JSON REFRESH) =======================================#
@login_required(login_url='staff_login')
def dashboard_api(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    search = request.GET.get("search", "")
    today = timezone.now().date()

    batches_base = Batch.objects.all()
    if my_role == 'Trainer':
        batches_base = batches_base.filter(trainer=my_staff)

    batches = filter_batches_by_search(batches_base, search)

    enrollments = Enrollment.objects.filter(batch__in=batches)
    attendance_qs = Attendance.objects.filter(enrollment__in=enrollments, attendance_date=today)

    counts = count_by_status(attendance_qs)

    batch_details = []

    for batch in batches:
        if Enrollment.objects.filter(batch=batch).exists():
            batch_details.append({
                "id": batch.id,
                "course": batch.course.course_name,
                "batch": batch.batch_name,
                "trainer": get_trainer_display_name(batch),
                "timing": get_batch_timing_display(batch),
                "session": batch.timing,
            })

    return JsonResponse({
        "total": enrollments.count(),
        "present": counts["present"],
        "absent": counts["absent"],
        "late": counts["late"],
        "batches": batch_details,
    })

#=========================================== DASHBOARD PAGE =====================================#
@login_required(login_url='staff_login')
def dashboard(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    search = request.GET.get("search", "")
    selected_date_param = request.GET.get("date")

    if selected_date_param:
        try:
            today = datetime.strptime(selected_date_param, "%Y-%m-%d").date()
        except ValueError:
            today = timezone.now().date()
    else:
        today = timezone.now().date()

    batches_base = Batch.objects.all()

    # Trainer -> only own assigned batches
    if my_role == 'Trainer':
        batches_base = batches_base.filter(trainer=my_staff)

    batches = filter_batches_by_search(batches_base, search)

    enrollments = Enrollment.objects.filter(batch__in=batches)
    total = enrollments.count()
    attendance_qs = Attendance.objects.filter(enrollment__in=enrollments, attendance_date=today)
    real_today = timezone.now().date()

    # Current Month
    current_month_students = Enrollment.objects.filter(
        batch__in=batches_base, start_date__year=real_today.year, start_date__month=real_today.month
    ).count()

    # Previous Month
    if real_today.month == 1:
        prev_month = 12
        prev_year = real_today.year - 1
    else:
        prev_month = real_today.month - 1
        prev_year = real_today.year

    previous_month_students = Enrollment.objects.filter(
        batch__in=batches_base, start_date__year=prev_year, start_date__month=prev_month
    ).count()

    # Percentage
    if previous_month_students > 0:
        total_percentage = round(
            ((current_month_students - previous_month_students) / previous_month_students) * 100, 2
        )
    else:
        total_percentage = 100 if current_month_students > 0 else 0

    today_counts = count_by_status(attendance_qs)
    present = today_counts["present"]
    absent = today_counts["absent"]
    late = today_counts["late"]

    attendance_marked = attendance_qs.count()

    last_attendance_date = Attendance.objects.filter(enrollment__in=enrollments, attendance_date__lt=today).order_by('-attendance_date').values_list(
        'attendance_date',
        flat=True
    ).first()

    last_present = 0
    last_absent = 0
    last_late = 0

    if last_attendance_date:
        last_qs = Attendance.objects.filter(
            enrollment__in=enrollments,
            attendance_date=last_attendance_date
        )
        last_counts = count_by_status(last_qs)
        last_present = last_counts["present"]
        last_absent = last_counts["absent"]
        last_late = last_counts["late"]

    percentage = round((present / total) * 100, 2) if total else 0

    present_percentage = round((present / total) * 100, 2) if total else 0
    absent_percentage = round((absent / total) * 100, 2) if total else 0
    late_percentage = round((late / total) * 100, 2) if total else 0

    last_total = last_present + last_absent + last_late

    last_present_percentage = round(
        (last_present / last_total) * 100, 2
    ) if last_total else 0

    last_absent_percentage = round(
        (last_absent / last_total) * 100, 2
    ) if last_total else 0

    last_late_percentage = round(
        (last_late / last_total) * 100, 2
    ) if last_total else 0

    present_change = round(present_percentage - last_present_percentage, 2)
    absent_change = round(absent_percentage - last_absent_percentage, 2)
    late_change = round(late_percentage - last_late_percentage, 2)

    batch_details = []

    for batch in batches:

        batch_enrollment_count = Enrollment.objects.filter(batch=batch).count()
        if batch_enrollment_count > 0:

            batch_today_attendance = Attendance.objects.filter(batch=batch, attendance_date=today)
            batch_counts = count_by_status(batch_today_attendance)

            batch_present = batch_counts["present"]
            batch_absent = batch_counts["absent"]
            batch_late = batch_counts["late"]

            batch_attendance_rate = round(
                ((batch_present + batch_late) / batch_enrollment_count) * 100, 1
            ) if batch_enrollment_count else 0

            batch_details.append({
                "id": batch.id,
                "course": batch.course.course_name,
                "batch": batch.batch_name,
                "trainer": get_trainer_display_name(batch),
                "timing": get_batch_timing_display(batch),
                "students": batch_enrollment_count,
                "present": batch_present,
                "absent": batch_absent,
                "late": batch_late,
                "attendance_rate": batch_attendance_rate,
            })

    available_dates_qs = Attendance.objects.filter(enrollment__in=enrollments).values_list('attendance_date', flat=True).distinct()
    available_dates_json = json.dumps([d.strftime('%Y-%m-%d') for d in available_dates_qs])

    context = {
        "present": present,
        "absent": absent,
        "late": late,
        "total": total,
        "percentage": percentage,
        "attendance_marked": attendance_marked,
        "present_change": present_change,
        "absent_change": absent_change,
        "late_change": late_change,
        "total_percentage": total_percentage,
        "context_batches": batch_details,
        "present_percentage": present_percentage,
        "absent_percentage": absent_percentage,
        "late_percentage": late_percentage,
        "last_attendance_date": last_attendance_date,
        "selected_date": today,
        "available_dates_json": available_dates_json,
    }

    return render(request, 'attendance/dashboard.html', context)

#==================================== BATCHES PAGE ======================================#
def batches_page(request):
    form = BatchForm()
    context = {'form': form}
    return render(request, 'attendance/batches.html', context)

#===================================== BATCH PREVIEW PAGE ================================#
@login_required(login_url='staff_login')
def batch_preview(request, batch_id):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    batch = get_object_or_404(Batch, id=batch_id)

    if my_role == 'Trainer' and batch.trainer_id != my_staff.id:
        messages.error(request, "You are not assigned to this batch.")
        return redirect('staff_dashboard')

    total_students = Enrollment.objects.filter(batch=batch).count()

    today = timezone.now().date()

    today_attendance = Attendance.objects.filter(batch=batch, attendance_date=today)
    counts = count_by_status(today_attendance)

    present_count = counts["present"]
    absent_count = counts["absent"]
    late_count = counts["late"]

    attendance_marked = today_attendance.exists()
    marked_total = counts["total"]

    attendance_percentage = round(
        ((present_count + late_count) / marked_total) * 100
    ) if marked_total else 0

    batch_code = 'BATCH-{}-{:03d}'.format(
        batch.created_at.strftime('%m%y'),
        batch.id
    )

    context = {
        'batch': batch,
        'batch_code': batch_code,
        'batch_status': batch.display_status,
        'total_students': total_students,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'attendance_marked': attendance_marked,
        'attendance_percentage': attendance_percentage,
        'today': today,
    }

    return render(request, 'attendance/batch_preview.html', context)

#===================================== MARK ATTENDANCE PAGE ======================================#
@login_required(login_url='staff_login')
def mark_attendance_page(request, batch_id):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    batch = get_object_or_404(Batch, id=batch_id)

    if my_role == 'Trainer' and batch.trainer_id != my_staff.id:
        messages.error(request, "You are not assigned to this batch.")
        return redirect('staff_dashboard')

    enrollments = Enrollment.objects.filter(batch=batch, admission__course_name=batch.course,)

    attendance_records = Attendance.objects.filter(batch=batch, attendance_date=timezone.now().date())

    syllabus_log = SyllabusLog.objects.filter(batch=batch, date=timezone.now().date()).first()

    duration = syllabus_log.duration if syllabus_log else 0

    attendance_map = {
        att.enrollment_id: att.status
        for att in attendance_records
    }

    remarks_map = {
        att.enrollment_id: att.remarks
        for att in attendance_records
    }

    context = {
        'batch': batch,
        'enrollments': enrollments,
        'attendance_map': attendance_map,
        'remarks_map': remarks_map,
        'syllabus_log': syllabus_log,
        "duration_hours": duration // 60,
        "duration_minutes": duration % 60,
    }

    return render(request, 'attendance/mark_attendance.html', context)

#=================================== SAVE BULK ATTENDANCE (API) ====================================#
@api_view(['POST'])
@transaction.atomic
def bulk_attendance(request):

    if not request.user.is_authenticated:
        return Response({
                'status': False,
                'message': 'Login required'
            }, status=401
        )

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    batch_id = request.data.get('batch')
    attendance_list = request.data.get('attendance', [])
    syllabus_data = request.data.get('syllabus_log', {})

    if not syllabus_data.get('topic_covered'):
        return Response({
                'status': False,
                'message': 'Topic covered is required'
            }, status=400
        )

    duration = syllabus_data.get('duration')

    if not duration:
        return Response({
                'status': False,
                'message': 'Duration is required'
            }, status=400
        )

    try:
        batch_obj = Batch.objects.get(id=batch_id)

    except Batch.DoesNotExist:
        return Response({
                'status': False,
                'message': 'Batch not found'
            }, status=404
        )

    if my_role == 'Trainer' and batch_obj.trainer_id != my_staff.id:
        return Response({
                'status': False,
                'message': 'You are not assigned to this batch'
            }, status=403
        )

    attendance_date = timezone.now().date()

    for item in attendance_list:

        Attendance.objects.update_or_create(enrollment_id=item['enrollment'], batch_id=batch_id, attendance_date=attendance_date,
                                            defaults={
                                                'status': item['status'],
                                                'remarks': item.get('remarks', ''),
                                                'trainer': batch_obj.trainer
                                                })
        SyllabusLog.objects.update_or_create(batch=batch_obj, date=attendance_date,
                                             defaults={
                                                'trainer': batch_obj.trainer,
                                                'topic_covered': syllabus_data.get('topic_covered'),
                                                'duration': syllabus_data.get('duration'),
                                                'next_topic': syllabus_data.get('next_topic', ''),
                                                'trainer_notes': syllabus_data.get('trainer_notes', '')
                                                })

    return Response({
        'status': True,
        'message': 'Attendance and syllabus log saved successfully'
    })

#======================================= ATTENDANCE HISTORY PAGE =====================================#
@login_required(login_url='staff_login')
def attendance_history_page(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    records = Attendance.objects.select_related('enrollment', 'batch').order_by('-attendance_date')

    # Trainer -> only own assigned batches
    if my_role == 'Trainer':
        records = records.filter(batch__trainer=my_staff)

    courses = Course.objects.all()

    attendance_filter = AttendanceFilter(request.GET, queryset=records)

    # ---- Restrict batch dropdown to trainer's own batches ----
    if my_role == 'Trainer':
        attendance_filter.form.fields['batch'].queryset = Batch.objects.filter(trainer=my_staff)

    filtered_records = attendance_filter.qs

    total_records = filtered_records.count()
    present = filtered_records.filter(status__in=["Present", "Late"]).count()

    counts = count_by_status(filtered_records)
    total_present = counts["present"]
    total_absent = counts["absent"]
    total_late = counts["late"]

    avg_attendance = round(
        ((total_present + total_late) / total_records) * 100,
        2
    ) if total_records else 0

    context = {
        "filter": attendance_filter,
        "records": filtered_records,
        "courses": courses,
        "present": present,
        "total_records": total_records,
        "total_present": total_present,
        "total_absent": total_absent,
        "total_late": total_late,
        "avg_attendance": avg_attendance,
    }

    return render(request, 'attendance/attendance_history.html', context)

#=============================== ATTENDANCE EXPORT (EXCEL / PDF) ==================================#
@login_required(login_url='staff_login')
def attendance_export(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    records = Attendance.objects.select_related(
        'enrollment', 'batch', 'trainer'
    ).order_by('-attendance_date')

    if my_role == 'Trainer':
        records = records.filter(batch__trainer=my_staff)

    attendance_filter = AttendanceFilter(request.GET, queryset=records)
    qs = attendance_filter.qs

    # ---- Restrict to selected rows if any were checked ----
    selected_ids = request.GET.get("selected_ids", "")

    if selected_ids:

        id_list = [
            int(i) for i in selected_ids.split(",")
            if i.strip().isdigit()
        ]

        if id_list:
            qs = qs.filter(id__in=id_list)

    export_format = request.GET.get("format")
    export_rows = build_export_rows(qs)

    # ================= EXCEL =================
    if export_format == "excel":

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        headers = ['Date', 'Student', 'Course', 'Batch', 'Status', 'Trainer']
        ws.append(headers)

        header_fill = PatternFill(
            start_color="FFC000",
            end_color="FFC000",
            fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in export_rows:
            ws.append([
                row["date"], row["student_name"], row["course_name"],
                row["batch_name"], row["status"], row["trainer_name"]
            ])

        column_widths = {'A': 20, 'B': 25, 'C': 20, 'D': 18, 'E': 15, 'F': 20}

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'

        wb.save(response)
        return response

    # ================= PDF =================
    elif export_format == "pdf":

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)

        styles = getSampleStyleSheet()
        title = Paragraph("Attendance Report", styles['Title'])

        data = [['Date', 'Student', 'Course', 'Batch', 'Status', 'Trainer']]

        for row in export_rows:
            data.append([
                row["date"], row["student_name"], row["course_name"],
                row["batch_name"], row["status"], row["trainer_name"]
            ])

        table = Table(data)

        table.setStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.gold),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),

            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),

            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ])

        elements = [title, Spacer(1, 12), table]

        doc.build(elements)
        return response

    return HttpResponse("Invalid format", status=400)

#=============================== STUDENT ATTENDANCE SUMMARY (API) ===============================#
def student_attendance_summary(request, student_id):

    records = Attendance.objects.filter(enrollment__admission__student_id=student_id).order_by("-attendance_date")

    counts = count_by_status(records)
    present = counts["present"]
    absent = counts["absent"]
    late = counts["late"]

    total = records.count()

    percentage = round(((present + late) / total) * 100, 2) if total else 0

    student = records.first().enrollment.student if records.exists() else None

    timeline = []

    first_record = records.first()

    for r in records:
        timeline.append({
            "date": r.attendance_date.strftime("%d-%m-%Y"),
            "status": r.status
        })

    context = {
        "student_id": f"STU{student.id}",
        "course": first_record.enrollment.course.course_name if first_record else "",
        "batch": first_record.batch.batch_name if first_record else "",
        "timing": first_record.batch.timing if first_record else "",
        "photo_url": student.photo.url if student and student.photo else None,
        "present": present,
        "absent": absent,
        "late": late,
        "percentage": percentage,
        "timeline": timeline
    }

    return JsonResponse({"student_name": (
            f"{student.first_name} {student.last_name}"
            if student else ""
        ), **context
    })


#=================================== ABSENT TRACKER PAGE ====================================#
@login_required(login_url='staff_login')
def absent_tracker(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    absent_students = get_absent_tracker_data()

    absent_students = [
        student
        for student in absent_students
        if student['total_absences'] > 0
    ]

    # Trainer -> only students from own assigned batches
    if my_role == 'Trainer':
        absent_students = [
            student
            for student in absent_students
            if student['batch'] and student['batch'].trainer_id == my_staff.id
        ]

    absent_students = sorted(
        absent_students,
        key=lambda x: (x['consecutive_absences'], x['total_absences']),
        reverse=True
    )

    courses = Course.objects.all()
    batches = Batch.objects.all()

    if my_role == 'Trainer':
        batches = batches.filter(trainer=my_staff)

    total_students = len(absent_students)

    attention_students = len([
       student
       for student in absent_students
       if student['consecutive_absences'] >= 3
    ])

    if total_students > 0:
      attention_percentage = round(
        (attention_students / total_students) * 100
    )
    else:
      attention_percentage = 0

    active_alerts = attention_students

    if absent_students:
        max_consecutive_absence = max(
         student['consecutive_absences']
          for student in absent_students
    )
    else:
        max_consecutive_absence = 0

    at_risk_students = len([
       student
       for student in absent_students
       if student['consecutive_absences'] >= 3
    ])

    if total_students > 0:
       at_risk_percentage = round((at_risk_students / total_students) * 100)
    else:
       at_risk_percentage = 0

    context = {
        'absent_students': absent_students,
        'courses': courses,
        'batches': batches,
        'active_alerts': active_alerts,
        'total_students': total_students,
        'max_consecutive_absence': max_consecutive_absence,
        'at_risk_percentage': at_risk_percentage,
        'attention_students': attention_students,
        'attention_percentage': attention_percentage,
    }
    return render(request, 'attendance/absent_tracker.html', context)

#=========================== ABSENT TRACKER - NOTIFICATION STATUS ================================#
def mark_notification_sent(enrollment_id):

    tracker = AbsentTracker.objects.get_or_create(enrollment_id=enrollment_id)

    tracker.notification_sent = True
    tracker.notification_status = "Dispatched"
    tracker.last_notified_at = timezone.now()
    tracker.save()

    return JsonResponse({"status": "success"})

#================================== ABSENT TRACKER - ADMIN NOTES ====================================#
def get_admin_notes(tracker_id):

    tracker = AbsentTracker.objects.filter(
        id=tracker_id
    ).first()

    return JsonResponse({
        "notes": tracker.admin_notes
        if tracker and tracker.admin_notes
        else ""
    })


@require_POST
def save_admin_notes(request):

    data = json.loads(request.body)

    tracker_id = data.get("tracker_id")

    if not tracker_id or tracker_id == "None":
        return JsonResponse({"status": "error", "message": "Tracker ID not found"}, status=400)

    tracker = AbsentTracker.objects.filter(id=int(tracker_id)).first()

    if not tracker:
        return JsonResponse({"status": "error", "message": "Tracker not found"})

    notes = data.get("notes", "")

    tracker.admin_notes = notes
    tracker.save()

    return JsonResponse({"status": "success"})

#==================================== LOW ATTENDANCE ALERTS PAGE =================================#
@login_required(login_url='staff_login')
def low_attendance_alerts(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    low_attendance_students = (get_low_attendance_data())

    # Trainer -> only students from own assigned batches
    if my_role == 'Trainer':
        low_attendance_students = [
            student for student in low_attendance_students
            if student['batch'] and student['batch'].trainer_id == my_staff.id
        ]

    name_filter = request.GET.get("name")

    if name_filter:
        low_attendance_students = [
            student for student in low_attendance_students
            if name_filter.lower() in (
                f"{student['student'].first_name} {student['student'].last_name}".lower()
            )
        ]

    course_filter = request.GET.get("course")

    if course_filter:
        low_attendance_students = [
            student for student in low_attendance_students
            if student["course"].course_name.lower() == course_filter.lower()
        ]

    batch_filter = request.GET.get("batch")

    if batch_filter:
        low_attendance_students = [
            student for student in low_attendance_students
            if str(student["batch"].id) == str(batch_filter)
        ]

    attendance_filter = request.GET.get("attendance")

    if attendance_filter:
        low_attendance_students = [
            student
            for student in low_attendance_students
            if student["attendance_percentage"] == float(attendance_filter)
        ]

    critical_students = [
        student
        for student in low_attendance_students
        if student["alert_level"] == "Critical"
    ]

    warning_students = [
        student
        for student in low_attendance_students
        if student["alert_level"] == "Warning"
    ]

    total_students = len(low_attendance_students)

    overall_average = (
        round(sum(s["attendance_percentage"] for s in low_attendance_students) / total_students, 1)
        if total_students > 0
        else 0
    )

    courses = Course.objects.all()
    batches = Batch.objects.all()

    if my_role == 'Trainer':
        batches = batches.filter(trainer=my_staff)

    context = {
        'low_attendance_students': low_attendance_students,
        "critical_students": critical_students,
        "warning_students": warning_students,
        "total_students": total_students,
        "overall_average": overall_average,
        'courses': courses,
        'batches': batches,
        "filters": {
                "name": name_filter or "",
                "course": course_filter or "",
                "batch": batch_filter or "",
                "attendance": attendance_filter or "",
            },
    }

    return render(request, 'attendance/low_attendance.html', context)

#=============================== LOW ATTENDANCE EXPORT (EXCEL) ====================================#
@login_required(login_url='staff_login')
def low_attendance_export(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    alert_type = request.GET.get("type")

    students = get_low_attendance_data()

    if my_role == 'Trainer':
        students = [
            s for s in students
            if s['batch'] and s['batch'].trainer_id == my_staff.id
        ]

    if alert_type == "critical":
        students = [s for s in students if s["alert_level"] == "Critical"]
        file_name = "critical_alerts.xlsx"
    else:
        students = [s for s in students if s["alert_level"] == "Warning"]
        file_name = "warning_alerts.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Low Attendance"

    headers = ["Student Name", "Course", "Batch", "Attendance %", "Consecutive Absences", "Total Absences"]

    ws.append(headers)

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for student in students:

        ws.append([
            f"{student['student'].first_name} {student['student'].last_name}",
            student['course'].course_name,
            student['batch'].batch_name,
            student['attendance_percentage'],
            student['consecutive_absences'],
            student['total_absences']
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = (f'attachment; filename="{file_name}"')

    wb.save(response)

    return response

#============================= NOTIFICATIONS - SINGLE STUDENT (EMAIL) ============================#
def send_low_attendance_email(request, enrollment_id):

    enrollment = get_object_or_404(Enrollment, id=enrollment_id)
    student = enrollment.student

    attendance_percentage = get_student_attendance_percentage(enrollment)
    subject, message = build_attendance_email(student, attendance_percentage)

    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [student.email], fail_silently=False)

    messages.success(request, f"Email sent to {student.email}")

    return JsonResponse({"status": "success", "message": "📧 Email sent successfully"})

#============================ NOTIFICATIONS - SINGLE STUDENT (SMS) ================================#
def send_sms_notification(request, enrollment_id):

    enrollment = get_object_or_404(Enrollment, id=enrollment_id)
    student = enrollment.student

    attendance_percentage = get_student_attendance_percentage(enrollment)
    sms_message = build_attendance_sms(student, attendance_percentage)

    messages.success(request, f"SMS sent to {student.phone_no}")

    return JsonResponse({"status": "success", "message": "📱 SMS sent successfully"})

#===================== NOTIFICATIONS - BULK (SELECTED STUDENTS) ========================#
@require_POST
def send_bulk_notification(request):
    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({"message": "Invalid request data"}, status=400)

    notification_type = payload.get("type")
    enrollment_ids = payload.get("enrollment_ids", [])

    if notification_type not in ("sms", "email"):
        return JsonResponse({"message": "Invalid notification type"}, status=400)

    if not enrollment_ids:
        return JsonResponse({"message": "No students selected"}, status=400)

    sent_count = 0

    for enrollment_id in enrollment_ids:

        enrollment = Enrollment.objects.filter(id=enrollment_id).first()

        if not enrollment:
            continue

        student = enrollment.student
        attendance_percentage = get_student_attendance_percentage(enrollment)

        if notification_type == "email":

            if not student.email:
                continue

            subject, message = build_attendance_email(student, attendance_percentage)
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [student.email], fail_silently=False)
            sent_count += 1

        else:
            build_attendance_sms(student, attendance_percentage)
            sent_count += 1

    label = "Emails" if notification_type == "email" else "SMS"

    return JsonResponse({"status": "success", "message": f"✅ {label} sent to {sent_count} student(s)"})

#====================== NOTIFICATIONS - EMAIL ALL LOW ATTENDANCE STUDENTS =======================#
def send_email_all():

    low_attendance_students = (get_low_attendance_data())

    for student_data in low_attendance_students:
        student = student_data["student"]
        attendance_percentage = (student_data["attendance_percentage"])

        subject, message = build_attendance_email(student, attendance_percentage)

        if student.email:
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [student.email], fail_silently=False)

    return JsonResponse({"status": "success","message": "All Emails sent successfully"})

#===================== NOTIFICATIONS - SMS ALL LOW ATTENDANCE STUDENTS ======================#
def send_sms_all(request):

    low_attendance_students = (get_low_attendance_data())

    for student_data in low_attendance_students:
        student = student_data["student"]
        attendance_percentage = (student_data["attendance_percentage"])

        build_attendance_sms(student, attendance_percentage)

    messages.success(request, "SMS notifications sent.")

    return JsonResponse({"status": "success","message": " All SMS notifications sent"})

#================================ MONTHLY REPORT EMAIL ======================================#
def send_monthly_report(request):

    low_attendance_students = (get_low_attendance_data())

    report_lines = []

    report_lines.append("Monthly Low Attendance Report\n")

    for student_data in low_attendance_students:
        report_lines.append(f"""
        Student:{student_data['student'].first_name}
        {student_data['student'].last_name}

        Course:{student_data['course'].course_name}

        Batch:{student_data['batch'].batch_name}

        Attendance:{student_data['attendance_percentage']}%

        Total Absences:{student_data['total_absences']}""")

    report_content = "\n".join(report_lines)

    send_mail("Monthly Attendance Report", report_content, settings.DEFAULT_FROM_EMAIL, [settings.DEFAULT_FROM_EMAIL], fail_silently=False)

    messages.success(request, "Monthly report sent.")

    return JsonResponse({"status": "success","message": "📊 Monthly report sent successfully"})


#================================================================
# REPORTS & ANALYTICS SECTION
# (report table data + PDF/Excel exports for Reports page)
#================================================================

def get_report_students(my_staff=None, my_role=None):
    report_students = []
    enrollments = Enrollment.objects.select_related('admission__student', 'admission__course_name', 'batch')

    if my_role == 'Trainer' and my_staff:
        enrollments = enrollments.filter(batch__trainer=my_staff)

    for enrollment in enrollments:
        present_count, absent_count, late_count, total_days = get_attendance_counts(enrollment)

        if total_days == 0:
            continue

        status, attendance_rate = get_attendance_status(present_count, absent_count, late_count, total_days)

        report_students.append({
            "student": enrollment.admission.student,
            "course": enrollment.admission.course_name,
            "batch": enrollment.batch,
            "present_count": present_count,
            "absent_count": absent_count,
            "late_count": late_count,
            "attendance_rate": attendance_rate,
            "status": status,
            "total_days": total_days,
        })

    return report_students

#=========================== REPORTS PAGE (MAIN VIEW + FILTERS) ==============================#
@login_required(login_url='staff_login')
def reports(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    today = timezone.now().date()
    attendance_date = get_formatted_latest_attendance_date()

    student_name = request.GET.get("student_name")
    attendance_filter = request.GET.get("attendance")
    course_filter = request.GET.get("course")
    batch_filter = request.GET.get("batch")
    status_filter = request.GET.get("status")
    date_from_filter = request.GET.get("date_from")
    date_to_filter = request.GET.get("date_to")
    distribution_percentage_filter = request.GET.get("distribution_percentage")

    validation_errors = []

    trainer_batch_ids = None
    if my_role == 'Trainer':
        trainer_batch_ids = list(Batch.objects.filter(trainer=my_staff).values_list('id', flat=True))

    if batch_filter:
        try:
            batch_filter_id = int(batch_filter)
        except (TypeError, ValueError):
            validation_errors.append("Selected batch value is invalid, so the batch filter was ignored.")
            batch_filter = None
        else:
            batch_lookup = Batch.objects.filter(id=batch_filter_id)
            if my_role == 'Trainer':
                batch_lookup = batch_lookup.filter(trainer=my_staff)
            if not batch_lookup.exists():
                validation_errors.append("Selected batch could not be found, so the batch filter was ignored.")
                batch_filter = None

    if course_filter:
        if not Course.objects.filter(course_name__iexact=course_filter).exists():
            validation_errors.append("Selected course could not be found, so the course filter was ignored.")
            course_filter = None

    valid_status_values = {"excellent", "good", "warning", "critical"}
    if status_filter and status_filter.lower() not in valid_status_values:
        validation_errors.append("Selected status is not recognized, so the status filter was ignored.")
        status_filter = None

    if attendance_filter:
        try:
            attendance_filter_value = float(attendance_filter)
            if attendance_filter_value < 0 or attendance_filter_value > 100:
                raise ValueError
        except (TypeError, ValueError):
            validation_errors.append("Attendance % must be a number between 0 and 100, so the attendance filter was ignored.")
            attendance_filter = None

    if distribution_percentage_filter:
        try:
            distribution_percentage_value = float(distribution_percentage_filter)
            if distribution_percentage_value < 0 or distribution_percentage_value > 100:
                raise ValueError
        except (TypeError, ValueError):
            validation_errors.append("Distribution % must be a number between 0 and 100, so that filter was ignored.")
            distribution_percentage_filter = None

    def _parse_filter_date(value):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None

    parsed_date_from = _parse_filter_date(date_from_filter) if date_from_filter else None
    parsed_date_to = _parse_filter_date(date_to_filter) if date_to_filter else None

    if date_from_filter and not parsed_date_from:
        validation_errors.append("From date is invalid, so the date range filter was ignored.")
        date_from_filter = None

    if date_to_filter and not parsed_date_to:
        validation_errors.append("To date is invalid, so the date range filter was ignored.")
        date_to_filter = None

    if parsed_date_from and parsed_date_to and parsed_date_from > parsed_date_to:
        validation_errors.append("From date cannot be after To date, so the date range filter was ignored.")
        date_from_filter = None
        date_to_filter = None

    batch_chart_title = "Batch-wise Attendance"

    if batch_filter:
        selected_batch_obj = Batch.objects.filter(id=batch_filter).first()
        if selected_batch_obj:
            batch_chart_title = (
                f"Batch-wise Attendance - "
                f"{selected_batch_obj.batch_name}"
                f"({selected_batch_obj.course.course_name})"
            )
    elif course_filter:
        batch_chart_title = f"Batch-wise Attendance - {course_filter.title()}"

    total_students_qs = Enrollment.objects.all()
    attendance_today_qs = Attendance.objects.filter(attendance_date=today)

    if my_role == 'Trainer':
        total_students_qs = total_students_qs.filter(batch__trainer=my_staff)
        attendance_today_qs = attendance_today_qs.filter(batch__trainer=my_staff)

    total_students = total_students_qs.count()
    today_marked_count = attendance_today_qs.count()
    pending_count = total_students - today_marked_count

    if today_marked_count == 0:
        attendance_status = "not_started"
    elif pending_count > 0:
        attendance_status = "in_progress"
    else:
        attendance_status = "completed"

    present_today = attendance_today_qs.filter(Q(status='Present') | Q(status='Late')).count()
    absent_today = attendance_today_qs.filter(status='Absent').count()

    low_attendance_all = get_low_attendance_data()
    if my_role == 'Trainer':
        low_attendance_all = [
            s for s in low_attendance_all
            if s['batch'] and s['batch'].trainer_id == my_staff.id
        ]
    low_attendance = len(low_attendance_all)

    report_students = []

    enrollments = Enrollment.objects.select_related('admission__student', 'admission__course_name', 'batch')

    if my_role == 'Trainer':
        enrollments = enrollments.filter(batch__trainer=my_staff)

    if course_filter:
        enrollments = enrollments.filter(admission__course_name__course_name__iexact=course_filter)
    if batch_filter:
        enrollments = enrollments.filter(batch_id=batch_filter)
    if student_name:
        enrollments = enrollments.filter(admission__student__first_name__istartswith=student_name)

    for enrollment in enrollments:
        present_count, absent_count, late_count, total_days = get_attendance_counts(
            enrollment, date_from_filter, date_to_filter
        )

        if total_days == 0:
            continue

        status, attendance_rate = get_attendance_status(present_count, absent_count, late_count, total_days)

        if attendance_filter and round(attendance_rate, 1) != round(float(attendance_filter), 1):
            continue

        if status_filter and status.lower() != status_filter.lower():
            continue

        report_students.append({
            "enrollment_id": enrollment.id,
            "student": enrollment.admission.student,
            "course": enrollment.admission.course_name,
            "batch": enrollment.batch,
            "present_count": present_count,
            "absent_count": absent_count,
            "late_count": late_count,
            "attendance_rate": attendance_rate,
            "status": status,
            "total_days": total_days,
        })

    filtered_enrollment_ids = [s["enrollment_id"] for s in report_students]
    enrollments = enrollments.filter(id__in=filtered_enrollment_ids)

    monthly_chart_title = "Monthly Attendance Chart"
    monthly_present = [0] * 12
    monthly_absent = [0] * 12
    monthly_late = [0] * 12
    monthly_incomplete = [0] * 12

    attendance_qs = Attendance.objects.all()

    if my_role == 'Trainer':
        attendance_qs = attendance_qs.filter(batch__trainer=my_staff)

    if student_name:
        attendance_qs = attendance_qs.filter(enrollment__admission__student__first_name__istartswith=student_name)

    if status_filter or attendance_filter:
        attendance_qs = attendance_qs.filter(enrollment_id__in=filtered_enrollment_ids)

    if course_filter:
        attendance_qs = attendance_qs.filter(enrollment__admission__course_name__course_name__icontains=course_filter)

    if batch_filter:
        attendance_qs = attendance_qs.filter(enrollment__batch_id=batch_filter)

    if date_from_filter:
        attendance_qs = attendance_qs.filter(attendance_date__gte=date_from_filter)

    if date_to_filter:
        attendance_qs = attendance_qs.filter(attendance_date__lte=date_to_filter)

    for month in range(1, 13):
        monthly_present[month - 1] = attendance_qs.filter(
            attendance_date__month=month, attendance_date__year=today.year, status="Present"
        ).count()

        monthly_absent[month - 1] = attendance_qs.filter(
            attendance_date__month=month, attendance_date__year=today.year, status="Absent"
        ).count()

        monthly_late[month - 1] = attendance_qs.filter(
            attendance_date__month=month, attendance_date__year=today.year, status="Late"
        ).count()

        month_marked_days = attendance_qs.filter(
            attendance_date__month=month, attendance_date__year=today.year
        ).values("attendance_date").distinct().count()

        month_expected = enrollments.count() * month_marked_days

        monthly_incomplete[month - 1] = max(
            month_expected - (monthly_present[month - 1] + monthly_absent[month - 1] + monthly_late[month - 1]), 0
        )

    course_labels = []
    course_counts = []
    courses = Course.objects.all()

    for course in courses:
        course_labels.append(course.course_name)
        course_enrollment_qs = Enrollment.objects.filter(admission__course_name=course)
        if my_role == 'Trainer':
            course_enrollment_qs = course_enrollment_qs.filter(batch__trainer=my_staff)
        course_counts.append(course_enrollment_qs.count())

    batch_labels = []
    batch_counts = []
    batch_present_counts = []
    batch_performance_labels = []
    batch_performance_counts = []
    batch_present_list = []
    batch_absent_list = []
    batch_percentage_list = []
    batch_late_list = []

    batches = Batch.objects.all()

    if my_role == 'Trainer':
        batches = batches.filter(trainer=my_staff)

    if course_filter:
        batches = batches.filter(course__course_name__icontains=course_filter)

    if batch_filter:
        batches = batches.filter(id=batch_filter)

    for batch in batches:
        stats = compute_batch_stats(batch, enrollment_ids=filtered_enrollment_ids, fallback_to_latest=True)

        batch_labels.append(f"{batch.course.course_name} - {batch.batch_name}")
        batch_counts.append(Enrollment.objects.filter(batch=batch).count())

        batch_present_list.append(stats["present"])
        batch_absent_list.append(stats["absent"])
        batch_late_list.append(stats["late"])
        batch_percentage_list.append(stats["percentage"])
        batch_present_counts.append(stats["percentage"])
        batch_performance_labels.append(f"{batch.course.course_name} - {batch.batch_name}")
        batch_performance_counts.append(stats["percentage"])

    if course_filter:
        monthly_chart_title = f"Monthly Attendance Chart - {course_filter.title()}"

    is_filtered = bool(course_filter or batch_filter or student_name or status_filter or attendance_filter)
    total_attendance = sum(batch_present_list) + sum(batch_absent_list) + sum(batch_late_list)

    show_batch_chart = True
    if is_filtered and (course_filter or batch_filter):
        show_batch_chart = total_attendance > 0

    warning_message = ""
    if not show_batch_chart:
        if batch_filter:
            warning_message = "Attendance has not been marked for the selected batch today."
        elif course_filter:
            warning_message = "Attendance has not been marked for the selected course today."

    batch_status_map = {}
    for batch in batches:
        marked = Attendance.objects.filter(batch=batch, attendance_date=today).exists()
        batch_status_map[batch.id] = "not_started" if not marked else "done"

    distribution_date_from = date_from_filter or today
    distribution_date_to = date_to_filter or today

    qualifying_enrollment_ids = None
    if distribution_percentage_filter:
        try:
            min_percentage = float(distribution_percentage_filter)
            qualifying_enrollment_ids = [
                s["enrollment_id"] for s in report_students
                if s["attendance_rate"] >= min_percentage
            ]
        except ValueError:
            qualifying_enrollment_ids = None

    distribution_enrollments = enrollments
    if qualifying_enrollment_ids is not None:
        distribution_enrollments = enrollments.filter(id__in=qualifying_enrollment_ids)

    distribution_attendance_qs = Attendance.objects.filter(
        enrollment__in=distribution_enrollments,
        attendance_date__gte=distribution_date_from,
        attendance_date__lte=distribution_date_to,
    )

    distribution_present = distribution_attendance_qs.filter(status='Present').count()
    distribution_absent = distribution_attendance_qs.filter(status='Absent').count()
    distribution_late = distribution_attendance_qs.filter(status='Late').count()

    distribution_days_count = distribution_attendance_qs.values('attendance_date').distinct().count() or 1
    distribution_expected_days = distribution_enrollments.count() * distribution_days_count

    distribution_incomplete = max(
        distribution_expected_days - (distribution_present + distribution_absent + distribution_late), 0
    )

    distribution_total = distribution_present + distribution_absent + distribution_late + distribution_incomplete

    def _pct(part):
        return round((part / distribution_total) * 100, 1) if distribution_total else 0

    distribution_data = {
        "present": distribution_present,
        "absent": distribution_absent,
        "late": distribution_late,
        "incomplete": distribution_incomplete,
        "present_pct": _pct(distribution_present),
        "absent_pct": _pct(distribution_absent),
        "late_pct": _pct(distribution_late),
        "incomplete_pct": _pct(distribution_incomplete),
    }

    all_batches_qs = Batch.objects.all()
    if my_role == 'Trainer':
        all_batches_qs = all_batches_qs.filter(trainer=my_staff)

    all_batch_labels = []
    all_batch_present_list = []
    all_batch_absent_list = []
    all_batch_late_list = []
    all_batch_percentage_list = []
    all_batch_ids = []

    for batch in all_batches_qs:
        stats = compute_batch_stats(batch, date_from=date_from_filter, date_to=date_to_filter, fallback_to_latest=True)

        all_batch_ids.append(batch.id)
        all_batch_labels.append(f"{batch.course.course_name} - {batch.batch_name}")
        all_batch_present_list.append(stats["present"])
        all_batch_absent_list.append(stats["absent"])
        all_batch_late_list.append(stats["late"])
        all_batch_percentage_list.append(stats["percentage"])

    context = {
        "total_students": total_students,
        "present_today": present_today,
        "absent_today": absent_today,
        "low_attendance": low_attendance,
        "report_students": report_students,
        "monthly_present": monthly_present,
        "monthly_absent": monthly_absent,
        "monthly_late": monthly_late,
        "monthly_incomplete": monthly_incomplete,
        "course_labels": course_labels,
        "course_counts": course_counts,
        "batch_labels": batch_labels,
        "batch_counts": batch_counts,
        "batches": batches,
        "courses": courses,
        "batch_present_counts": batch_present_counts,
        "batch_performance_labels": batch_performance_labels,
        "batch_performance_counts": batch_performance_counts,
        "attendance_status": attendance_status,
        "today_marked_count": today_marked_count,
        "pending_count": pending_count,
        "batch_present_list": batch_present_list,
        "batch_absent_list": batch_absent_list,
        "batch_percentage_list": batch_percentage_list,
        "batch_late_list": batch_late_list,
        "batch_chart_title": batch_chart_title,
        "monthly_chart_title": monthly_chart_title,
        "attendance_date": attendance_date,
        "batch_status_map": batch_status_map,
        "is_filtered": is_filtered,
        "show_batch_chart": show_batch_chart,
        "warning_message": warning_message,
        "date_from_filter": date_from_filter or "",
        "date_to_filter": date_to_filter or "",
        "distribution_percentage_filter": distribution_percentage_filter or "",
        "distribution_data": distribution_data,
        "all_batch_labels": all_batch_labels,
        "all_batch_ids": all_batch_ids,
        "all_batch_present_list": all_batch_present_list,
        "all_batch_absent_list": all_batch_absent_list,
        "all_batch_late_list": all_batch_late_list,
        "all_batch_percentage_list": all_batch_percentage_list,
        "validation_errors": validation_errors,
    }

    if request.headers.get("x-requested-with") == "XMLHttpRequest":

        table_rows_html = render_to_string(
            "attendance/_report_table_rows.html",
            {"report_students": report_students},
            request=request
        )

        filter_message_html = render_to_string(
            "attendance/_filter_message.html",
            {
                "attendance_status": attendance_status,
                "today_marked_count": today_marked_count,
                "total_students": total_students,
                "pending_count": pending_count,
                "validation_errors": validation_errors,
            },
            request=request
        )

        return JsonResponse({
            "monthly_present": monthly_present,
            "monthly_absent": monthly_absent,
            "monthly_late": monthly_late,
            "monthly_incomplete": monthly_incomplete,
            "monthly_chart_title": monthly_chart_title,
            "distribution_data": distribution_data,
            "table_rows_html": table_rows_html,
            "record_count": len(report_students),
            "filter_message_html": filter_message_html,
            "validation_errors": validation_errors,
        })

    return render(request, "attendance/reports.html", context)
#============================= ANALYTICS EXPORT - PDF =====================================#
@login_required(login_url='staff_login')
def analytics_pdf(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=analytics_report.pdf'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("CSC Computer Education", styles['Title']))
    elements.append(Paragraph("Attendance Analytics Report", styles['Heading2']))
    elements.append(Spacer(1, 15))

    today = timezone.now().date()

    total_students_qs = Enrollment.objects.all()
    present_today_qs = Attendance.objects.filter(attendance_date=today, status='Present')
    absent_today_qs = Attendance.objects.filter(attendance_date=today, status='Absent')

    if my_role == 'Trainer':
        total_students_qs = total_students_qs.filter(batch__trainer=my_staff)
        present_today_qs = present_today_qs.filter(batch__trainer=my_staff)
        absent_today_qs = absent_today_qs.filter(batch__trainer=my_staff)

    total_students = total_students_qs.count()
    present_today = present_today_qs.count()
    absent_today = absent_today_qs.count()

    low_attendance_all = get_low_attendance_data()
    if my_role == 'Trainer':
        low_attendance_all = [
            s for s in low_attendance_all
            if s['batch'] and s['batch'].trainer_id == my_staff.id
        ]
    low_attendance = len(low_attendance_all)

    attendance_date = get_formatted_latest_attendance_date()

    summary_data = [
        ["Metric", "Value"],
        ["Total Students", total_students],
        ["Present Today", present_today],
        ["Absent Today", absent_today],
        ["Low Attendance", low_attendance],
        ["Last Attendance Updated", attendance_date],
    ]

    summary_table = Table(summary_data, colWidths=[220, 150])
    summary_table.setStyle(get_default_table_style(header_color="#1e40af", zebra=True))

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Course & Batch Analytics", styles['Heading2']))

    analytics_header = [["Course", "Batch", "Students", "Total Days", "Present", "Absent", "Late", "Attendance %"]]
    analytics_rows = get_batch_analytics_rows(my_staff=my_staff, my_role=my_role)
    analytics_rows = [[Paragraph(row[0], styles['BodyText'])] + row[1:] for row in analytics_rows]
    analytics_data = analytics_header + analytics_rows

    analytics_table = Table(analytics_data, colWidths=[140, 75, 50, 55, 45, 45, 40, 68])
    analytics_table.setStyle(get_default_table_style(header_color="#2563eb", zebra=True))

    elements.append(analytics_table)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Attendance Summary", styles['Heading2']))

    report_students = get_report_students(my_staff=my_staff, my_role=my_role)
    counts = get_status_summary_counts(report_students)

    attendance_summary = [
        ["Status", "Students"],
        ["Excellent", counts["Excellent"]],
        ["Good", counts["Good"]],
        ["Warning", counts["Warning"]],
        ["Critical", counts["Critical"]],
    ]

    attendance_summary_table = Table(attendance_summary, colWidths=[180, 120])
    summary_style = get_default_table_style(header_color="#059669", zebra=True)
    summary_style.add('BACKGROUND', (0, 1), (-1, 1), colors.lightgreen)
    summary_style.add('BACKGROUND', (0, 2), (-1, 2), colors.yellow)
    summary_style.add('BACKGROUND', (0, 3), (-1, 3), colors.salmon)
    attendance_summary_table.setStyle(summary_style)

    elements.append(attendance_summary_table)

    doc.build(elements)

    return response
#================================ ANALYTICS EXPORT - EXCEL ================================#
@login_required(login_url='staff_login')
def analytics_excel(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Analytics"

    today = timezone.now().date()

    total_students_qs = Enrollment.objects.all()
    present_today_qs = Attendance.objects.filter(attendance_date=today, status='Present')
    absent_today_qs = Attendance.objects.filter(attendance_date=today, status='Absent')

    if my_role == 'Trainer':
        total_students_qs = total_students_qs.filter(batch__trainer=my_staff)
        present_today_qs = present_today_qs.filter(batch__trainer=my_staff)
        absent_today_qs = absent_today_qs.filter(batch__trainer=my_staff)

    total_students = total_students_qs.count()
    present_today = present_today_qs.count()
    absent_today = absent_today_qs.count()

    low_attendance_all = get_low_attendance_data()
    if my_role == 'Trainer':
        low_attendance_all = [
            s for s in low_attendance_all
            if s['batch'] and s['batch'].trainer_id == my_staff.id
        ]
    low_attendance = len(low_attendance_all)

    attendance_date = get_formatted_latest_attendance_date()
    generated_date = timezone.now().strftime("%d %b %Y %I:%M %p")

    ws.merge_cells("A1:H1")
    ws["A1"] = "CSC Computer Education - Attendance Analytics Report"
    ws["A1"].fill = ExcelStyles.title_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = ExcelStyles.center

    ws.merge_cells("A3:B3")
    ws["A3"] = "Summary Dashboard"
    ws["A3"].fill = ExcelStyles.section_fill
    ws["A3"].font = ExcelStyles.white_font

    summary_data = [
        ["Metric", "Value"],
        ["Total Students", total_students],
        ["Present Today", present_today],
        ["Absent Today", absent_today],
        ["Low Attendance", low_attendance],
        ["Last Attendance Updated", attendance_date],
        ["Report Generated On", generated_date],
    ]

    row_num = 4
    for row in summary_data:
        ws.append(row)
        if row_num == 4:
            style_header_row(ws, row_num, len(row))
        else:
            for cell in ws[row_num]:
                cell.border = ExcelStyles.thin_border
        row_num += 1

    start_row = row_num + 2
    ws.merge_cells(f"A{start_row}:H{start_row}")
    ws[f"A{start_row}"] = "Course & Batch Analytics"
    ws[f"A{start_row}"].fill = ExcelStyles.section_fill
    ws[f"A{start_row}"].font = ExcelStyles.white_font

    analytics_header = ["Course", "Batch", "Students", "Total Days", "Present", "Absent", "Late", "Attendance %"]
    header_row = start_row + 1

    for col_num, value in enumerate(analytics_header, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = value
    style_header_row(ws, header_row, len(analytics_header))

    data_row = header_row + 1
    for row_data in get_batch_analytics_rows(my_staff=my_staff, my_role=my_role):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=data_row, column=col_num)
            cell.value = value
            cell.border = ExcelStyles.thin_border
            cell.alignment = ExcelStyles.center
        data_row += 1

    report_students = get_report_students(my_staff=my_staff, my_role=my_role)
    counts = get_status_summary_counts(report_students)

    status_row = data_row + 2
    ws.merge_cells(f"A{status_row}:B{status_row}")
    ws[f"A{status_row}"] = "Attendance Status Summary"
    ws[f"A{status_row}"].fill = ExcelStyles.section_fill
    ws[f"A{status_row}"].font = ExcelStyles.white_font

    summary_header_row = status_row + 1
    ws.cell(summary_header_row, 1).value = "Status"
    ws.cell(summary_header_row, 2).value = "Students"
    style_header_row(ws, summary_header_row, 2)

    status_data = [
        ["Excellent", counts["Excellent"]],
        ["Good", counts["Good"]],
        ["Warning", counts["Warning"]],
        ["Critical", counts["Critical"]],
    ]

    current_row = summary_header_row + 1
    for status, count in status_data:
        ws.cell(current_row, 1).value = status
        ws.cell(current_row, 2).value = count

        for col in range(1, 3):
            ws.cell(current_row, col).fill = STATUS_FILL_MAP[status]
            ws.cell(current_row, col).border = ExcelStyles.thin_border
            ws.cell(current_row, col).alignment = ExcelStyles.center

        current_row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=analytics_report.xlsx'

    wb.save(response)

    return response

#=============================== ATTENDANCE REPORT EXPORT - PDF ===============================#
@login_required(login_url='staff_login')
def report_pdf(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    students = get_report_students(my_staff=my_staff, my_role=my_role)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=attendance_report.pdf'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("CSC Computer Education", styles['Title']))
    elements.append(Paragraph("Student Attendance Report", styles['Heading2']))

    attendance_date = get_formatted_latest_attendance_date()
    elements.append(Paragraph(f"Last Attendance Updated : {attendance_date}", styles['Normal']))
    elements.append(Spacer(1, 15))

    data = [["Student Name", "Course", "Batch", "Present", "Absent", "Late", "Attendance %", "Status"]]

    for student in students:
        data.append([
            f"{student['student'].first_name} {student['student'].last_name}",
            student['course'].course_name,
            student['batch'].batch_name if student['batch'] else "-",
            student['present_count'],
            student['absent_count'],
            student['late_count'],
            f"{student['attendance_rate']}%",
            student['status'],
        ])

    table = Table(data, colWidths=[90, 105, 65, 45, 45, 35, 75, 65])
    table_style = get_default_table_style(header_color="#1E40AF", zebra=True)

    def _rate_color_hex(rate):
        if rate == 100:
            return STATUS_COLOR_HEX["Excellent"]
        elif rate >= 75:
            return STATUS_COLOR_HEX["Good"]
        elif rate >= 60:
            return STATUS_COLOR_HEX["Warning"]
        return STATUS_COLOR_HEX["Critical"]

    for row_num, student in enumerate(students, start=1):
        table_style.add(
            'TEXTCOLOR', (7, row_num), (7, row_num),
            colors.HexColor(f"#{STATUS_COLOR_HEX[student['status']]}")
        )
        table_style.add(
            'TEXTCOLOR', (6, row_num), (6, row_num),
            colors.HexColor(f"#{_rate_color_hex(student['attendance_rate'])}")
        )

    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)

    return response

#=============================== ATTENDANCE REPORT EXPORT - EXCEL =============================#
@login_required(login_url='staff_login')
def report_excel(request):

    my_staff = getattr(request.user, 'staff_profile', None)
    my_role = my_staff.role.role_name if my_staff and my_staff.role else None

    students = get_report_students(my_staff=my_staff, my_role=my_role)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.merge_cells('A1:H1')
    ws['A1'] = "CSC Computer Education"
    ws['A1'].font = Font(size=18, bold=True, color="1E293B")
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    ws['A2'] = "Student Attendance Report"
    ws['A2'].font = Font(size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    attendance_date = get_formatted_latest_attendance_date()

    ws['A4'] = "Generated On"
    ws['B4'] = timezone.now().strftime("%d %b %Y")
    ws['D4'] = "Last Attendance Updated"
    ws['E4'] = attendance_date

    headers = ["Student Name", "Course", "Batch", "Present", "Absent", "Late", "Attendance %", "Status"]
    header_row = 6

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
    style_header_row(ws, header_row, len(headers))

    row_num = 7

    for student in students:
        ws.cell(row=row_num, column=1).value = f"{student['student'].first_name} {student['student'].last_name}"
        ws.cell(row=row_num, column=2).value = student['course'].course_name
        ws.cell(row=row_num, column=3).value = student['batch'].batch_name if student['batch'] else "-"
        ws.cell(row=row_num, column=4).value = student['present_count']
        ws.cell(row=row_num, column=5).value = student['absent_count']
        ws.cell(row=row_num, column=6).value = student['late_count']
        ws.cell(row=row_num, column=7).value = f"{student['attendance_rate']}%"
        ws.cell(row=row_num, column=8).value = student['status']

        if row_num % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor="F8FAFC")

        attendance_cell = ws.cell(row=row_num, column=7)
        rate = student['attendance_rate']

        if rate == 100:
            attendance_cell.font = Font(color=STATUS_COLOR_HEX["Excellent"], bold=True)
        elif rate >= 76:
            attendance_cell.font = Font(color=STATUS_COLOR_HEX["Good"], bold=True)
        elif rate >= 60:
            attendance_cell.font = Font(color=STATUS_COLOR_HEX["Warning"], bold=True)
        else:
            attendance_cell.font = Font(color=STATUS_COLOR_HEX["Critical"], bold=True)

        status_cell = ws.cell(row=row_num, column=8)
        status_cell.fill = PatternFill("solid", fgColor=STATUS_COLOR_HEX[student['status']])
        status_cell.font = Font(color="FFFFFF", bold=True)

        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = ExcelStyles.center
            cell.border = ExcelStyles.thin_border

        row_num += 1

    counts = get_status_summary_counts(students)

    summary_row = row_num + 2
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = "Attendance Summary"
    ws[f'A{summary_row}'].font = Font(bold=True, size=14)

    ws.cell(summary_row + 1, 1).value = "Excellent"
    ws.cell(summary_row + 1, 2).value = counts["Excellent"]

    ws.cell(summary_row + 2, 1).value = "Good"
    ws.cell(summary_row + 2, 2).value = counts["Good"]

    ws.cell(summary_row + 3, 1).value = "Warning"
    ws.cell(summary_row + 3, 2).value = counts["Warning"]

    ws.cell(summary_row + 4, 1).value = "Critical"
    ws.cell(summary_row + 4, 2).value = counts["Critical"]

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'

    wb.save(response)

    return response